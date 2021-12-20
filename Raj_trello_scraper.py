import requests
import datetime
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import pandas as pd
from gooey import GooeyParser, Gooey

production_board_idb = "601c6df669ba790ff0ba440a"


cards_links = "https://trello.com/1/cards"
labels_link = "https://api.trello.com/1/boards/{}/labels".format(production_board_idb)
lists_link = "https://api.trello.com/1/boards/{}/lists".format(production_board_idb)
all_cards_link = "https://api.trello.com/1/boards/{}/cards".format(production_board_idb)
users_link = "https://api.trello.com/1/boards/{}/members".format(production_board_idb)
apikey = '2e4a6967dcc93370d2582847d6a45726'
atoken = 'a0f7750db6dce905e830089813cb8f7c4611695c8264ba325b1266cb156e4274'
params_presented = {'key': apikey, 'token': atoken}


list_pd = pd.DataFrame.from_dict(data=requests.get(lists_link, params_presented).json(), orient="columns")
list_pd.set_index('id', inplace=True)
# cards_pd = pd.DataFrame.from_dict(data=requests.get(cards_links, params_presented).json(), orient="columns")
labels_pd = pd.DataFrame.from_dict(data=requests.get(labels_link, params_presented).json(), orient="columns")
labels_pd.set_index('id', inplace=True)
users_pd = pd.DataFrame.from_dict(data=requests.get(users_link, params_presented).json(), orient="columns")
users_pd.set_index('id', inplace=True)


# out_dir = r"C:\Users\user\Documents\test_area"


def organise_cards_in_list(list_id):
    cards_in_list = pd.DataFrame.from_dict(data=requests.get("https://api.trello.com/1/lists/{}/cards".format(list_id), params_presented).json(), orient="columns")
    headers = ['Project', 'Staff assigned', 'Deadline start', 'Deadline end', 'Comments', 'Description text']
    card_table_output = []
    for i, card in cards_in_list.iterrows():
        template = card.isTemplate
        if not template:
            labels_on_card = [a['name'] for a in card.labels]
            members_on_card = card.idMembers
            member_name_list = []
            if len(members_on_card) > 0:
                for member in members_on_card:
                    member_name = users_pd.loc[member, 'fullName']
                    member_name_list.append(member_name)
            try:
                due_end = datetime.datetime.strftime(datetime.datetime.strptime(card.due, '%Y-%m-%dT%H:%M:%S.%fZ'), '%d/%m/%Y')
            except TypeError:
                due_end = ''
            try:
                due_start = datetime.datetime.strftime(datetime.datetime.strptime(card.start, '%Y-%m-%dT%H:%M:%S.%fZ'), '%d/%m/%Y')
            except TypeError:
                due_start = ''
            name = card['name']
            desc_text = [a for a in card.desc.split('\n')[1:] if len(a) > 0]

            return_list = [name, ', '.join(member_name_list), due_start, due_end, ', '.join(labels_on_card), ', '.join(desc_text)]
            card_table_output.append(return_list)
    this_list_pd = pd.DataFrame(data=card_table_output, columns=headers)
    return this_list_pd


def main_process(odir):
    today = datetime.datetime.now().strftime('%Y%m%d')
    ofile = os.path.join(odir, 'Production_Report_%s.xlsx' % today)
    wb = openpyxl.Workbook()
    col_as_letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    for i, row in list_pd.iterrows():
        list_name = row['name']
        if len(list_name) > 30:
            list_name = list_name.replace('Trello card to be archived (m', 'M').replace(')', '')
        cards_table = organise_cards_in_list(row.name)
        if not cards_table.empty:
            ws = wb.create_sheet(list_name)
            dims = {}
            for b, c in enumerate(cards_table.columns):
                max_len = cards_table[c].str.len().max()
                col_label_len = len(c)
                letter = col_as_letter[b]
                dims[letter] = max(max_len, col_label_len)
            for r in dataframe_to_rows(cards_table, index=False, header=True):
                ws.append(r)
            for cell in ws['A'] + ws[1]:
                cell.style = 'Pandas'
            for col, value in dims.items():
                ws.column_dimensions[col].width = value
        else:
            print('Nothing in column "%s"...' % list_name)
    del wb['Sheet']
    wb.save(ofile)
    print('Output saved to %s' % ofile)


@Gooey(program_name="Production Trello Scraper", use_legacy_titles=True, required_cols=1, default_size=(750, 500))
def goo():
    parser = GooeyParser(description='Trello to XLSX')
    parser.add_argument('output_folder', metavar='Output location for report', widget='DirChooser')
    return parser.parse_args()


if __name__ == '__main__':
    main_process(goo().output_folder)
